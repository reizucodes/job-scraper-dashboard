from __future__ import annotations

import io

from openpyxl import load_workbook

from app.exporters import XlsxExporter


def test_xlsx_exporter_writes_header_and_rows() -> None:
    exporter = XlsxExporter()
    rows = [{"title": "Vue Developer", "company": "Acme", "source_id": "1"}]

    artifact = exporter.export(rows=rows)

    assert artifact.file_extension == "xlsx"
    assert artifact.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(io.BytesIO(artifact.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "title"
    assert sheet.cell(row=2, column=1).value == "Vue Developer"
